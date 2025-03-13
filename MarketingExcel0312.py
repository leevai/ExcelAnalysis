import re
import sys
from datetime import datetime

import openpyxl
import time
import os
from openpyxl.styles import PatternFill, Alignment
from copy import copy

start_time = time.time()
# 获取可执行文件的路径
executable_path = sys.argv[0]
# 将可执行文件路径转换为绝对路径
current_directory = os.path.dirname(os.path.abspath(executable_path))
config_path = os.path.join(current_directory, 'excelConfig.txt')
print(f"开始获取:" + config_path + "路径下excelConfig中Excel文件")
with open(config_path) as f:
    excel_name_new = f.readline().strip()
    excel_name_old = f.readline().strip()
index = excel_name_new.index('.xlsx')
if not excel_name_new.endswith('.xlsx'):
    print("只支持.xlsx格式文件")
    exit()
if excel_name_new == excel_name_old:
    print("Excel文件名不能相同")
    exit()

excel_name_new = os.path.join(current_directory, excel_name_new)
excel_name_old = os.path.join(current_directory, excel_name_old)
# 打开或创建工作簿
workbook_a = openpyxl.load_workbook(excel_name_new)
worksheet_a = workbook_a.active
workbook_b = openpyxl.load_workbook(excel_name_old)
worksheet_b = workbook_b.active

# 加载工作簿并获取旧表格的表头
worksheet_b_header = workbook_b.active  # 获取包含表头的工作表
header_row = worksheet_b_header[1]  # 假设表头在第一行（索引为0）

# 获取新表格并清空前两行内容（假设第一行为已存在的表头，第二行为可能的数据）
workbook_a = openpyxl.load_workbook(excel_name_new)
worksheet_a = workbook_a.active

# 创建一个红色背景填充样式
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type='solid')

# a_data_dict = {int(row[0]): row[1:] for row in worksheet_a.iter_rows(min_row=2, values_only=True)}
# 将worksheet_a的值存储在字典中
a_data_dict = {}
for row in worksheet_a.iter_rows(min_row=2, values_only=True):
    key_a = row[0]
    if isinstance(key_a, str):
        try:
            key_a = int(row[0])
            a_data_dict[key_a] = row[1:]
        except ValueError:
            print(f"Unable to convert '{row[0]}' to an integer. Skipping this row.")
    else:
        a_data_dict[key_a] = row[1:]

# 将新表格的数据和样式存入字段中
b_data_styles_dict = {}
for row in worksheet_b.iter_rows(min_row=2):
    row_data = []
    for cell in row:
        # 获取单元格的值
        value = cell.value
        # 获取单元格的样式
        style = cell.style
        # 将值和样式一同保存
        row_data.append((value, style))
    # 将这一行的数据及其样式保存到字典中，键为第一列的值
    key_b = row[0].value
    if isinstance(key_b, str):
        try:
            key_b = int(row[0])
            b_data_styles_dict[key_b] = row_data[1:]
        except ValueError:
            print(f"表格"+excel_name_new+"中,项目编号:"+key_b+"Unable to convert to an integer. Skipping this row.")
    else:
        b_data_styles_dict[key_b] = row_data[1:]

print(f"完成路径下excel文件获取和数据整理")


def str_match(org_name):
    pattern_jh = "京沪大区"
    pattern_west = "西部大区"
    pattern_south = "南部大区"
    pattern_mid = "中部大区"
    if re.search(pattern_jh, org_name):
        return pattern_jh
    elif re.search(pattern_west, org_name):
        return pattern_west
    elif re.search(pattern_south, org_name):
        return pattern_south
    elif re.search(pattern_mid, org_name):
        return pattern_mid
    else:
        return org_name


print(f"开始分析Excel数据")
# 遍历存量表格A:假设ID列在第一列（即A列），且数据从第二行开始
for row_index, row_a in enumerate(worksheet_a.iter_rows(min_row=2, values_only=True)):
    id_value = row_a[0]
    # 判断id_value是否为整数,如果不是则转化为整数类型
    if isinstance(id_value, str):
        # 捕获转换异常
        try:
            id_value = int(id_value)
        except ValueError:
            print(f"数据异常:发现非数字的项目编号:" + str(id_value))
            continue
    if id_value in b_data_styles_dict:
        b_row_data_style = b_data_styles_dict[id_value]
        for col_index, value in enumerate(row_a[1:], start=1):  # 跳过ID列
            if col_index - 1 < 26:
                if value != b_row_data_style[col_index - 1][0]:  # 比较非ID列数据是否不同
                    cell_a = worksheet_a.cell(row=row_index + 2, column=col_index + 1)  # 从第2行开始，在第2列开始执行
                    cell_a.value = b_row_data_style[col_index - 1][0]
                    cell_a.fill = red_fill
            elif col_index == 27:
                # 从第2行开始，在第2列开始执行
                cell_a = worksheet_a.cell(row=row_index + 2, column=col_index + 1)
                # 根据P列提取上级部门到AB列
                cell_a.value = str_match(str(worksheet_a.cell(row=row_index + 2, column=16).value))
    else:
        print(f"数据异常:在"+excel_name_new+"中存在，在"+excel_name_old+"不存在的项目编号:" + str(id_value))

for row_index, row_b in enumerate(worksheet_b.iter_rows(min_row=2)):
    id_value_b = row_b[0].value
    # 判断id_value是否为整数,如果不是则转化为整数类型
    if isinstance(id_value_b, str):
        # 捕获转换异常
        try:
            id_value_b = int(id_value_b)
        except ValueError:
            print(f"数据异常:发现非数字的项目编号:" + str(id_value_b))
            continue
    if id_value_b not in a_data_dict:
        # 在worksheet_a中新增条目
        num_rows_a = worksheet_a.max_row
        for col_index, cell_value_b in enumerate(row_b):
            cell_a = worksheet_a.cell(row=num_rows_a + 1, column=col_index + 1)
            cell_a.value = cell_value_b.value
            if hasattr(cell_value_b, 'font') and cell_value_b.font is not None:
                new_font = copy(cell_value_b.font)
                cell_a.font = new_font
            if hasattr(cell_value_b, 'fill') and cell_value_b.fill is not None:
                new_fill = copy(cell_value_b.fill)
                cell_a.fill = new_fill
            if hasattr(cell_value_b, 'border') and cell_value_b.border is not None:
                new_border = copy(cell_value_b.border)
                cell_a.border = new_border
            if hasattr(cell_value_b, 'number_format') and cell_value_b.number_format is not None:
                new_format = copy(cell_value_b.number_format)
                cell_a.number_format = new_format
            if hasattr(cell_value_b, 'protection') and cell_value_b.protection is not None:
                new_protection = copy(cell_value_b.protection)
                cell_a.protection = new_protection
            if hasattr(cell_value_b, 'hyperlink') and cell_value_b.hyperlink is not None:
                new_hyperlink = copy(cell_value_b.hyperlink)
                cell_a.hyperlink = new_hyperlink

        # 将新增的项目编号标红
        center_alignment = Alignment(horizontal='center', vertical='center')
        cell_a = worksheet_a.cell(row=num_rows_a + 1, column=1)
        cell_a.fill = red_fill
        cell_a.alignment = center_alignment
        cell_a = worksheet_a.cell(row=num_rows_a + 1, column=28)
        # 根据P列提取上级部门到AB列
        cell_a.value = str_match(str(worksheet_a.cell(row=num_rows_a + 1, column=16).value))
print(f"结束分析Excel数据")

# 保存更改后的A表格
file_name_without_extension, file_extension = os.path.splitext(os.path.basename(excel_name_new))
current_datetime = datetime.now()
formatted_string = current_datetime.strftime('%Y%m%d%H%M')
out_put_excel_name = file_name_without_extension + '_分析完成_' + formatted_string + '.xlsx'
out_put_excel_name = os.path.join(current_directory, out_put_excel_name)
workbook_a.save(out_put_excel_name)
print(f"数据分析完成,新的Excel文件路径:", out_put_excel_name)

# 关闭工作簿
workbook_a.close()
workbook_b.close()
end_time = time.time()
execution_time = end_time - start_time
print(f"分析数据总耗时：{execution_time}秒")
